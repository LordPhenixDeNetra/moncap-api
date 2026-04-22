from __future__ import annotations

import os
import uuid
from io import BytesIO
from datetime import date, datetime, timezone

import pytest

from app.core.settings import get_settings
from app.core.security import hash_password, normalize_email
from app.models.adhesion import Adhesion
from app.models.enums import AdhesionStatus, AppRole, EngagementType, PaymentMode
from app.models.geo import Commune, Departement, Region
from app.repositories.users import UserRepository


async def test_admin_routes_forbidden_without_admin_role(client, db_session):
    repo = UserRepository(db_session)
    await repo.create_user(email=normalize_email("user@test.com"), password_hash=hash_password("Password123!"))
    await db_session.commit()

    r = await client.post("/api/v1/auth/login", json={"email": "user@test.com", "password": "Password123!"})
    assert r.status_code == 200
    token = r.json()["data"]["accessToken"]

    r2 = await client.get("/api/v1/admin/adhesions", headers={"Authorization": f"Bearer {token}"})
    assert r2.status_code == 403

    r3 = await client.get("/api/v1/admin/adhesions/lookup", headers={"Authorization": f"Bearer {token}"})
    assert r3.status_code == 403

    r4 = await client.patch(
        f"/api/v1/admin/adhesions/{uuid.uuid4()}/payment",
        json={"paiementConfirme": True},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert r4.status_code == 403


async def test_admin_can_lookup_adhesion_details(client, db_session):
    repo = UserRepository(db_session)
    user = await repo.create_user(email=normalize_email("admin@test.com"), password_hash=hash_password("Password123!"))
    await repo.add_role(user_id=user.id, role=AppRole.admin)
    await db_session.commit()

    r = await client.post("/api/v1/auth/login", json={"email": "admin@test.com", "password": "Password123!"})
    assert r.status_code == 200
    token = r.json()["data"]["accessToken"]

    region = Region(id=uuid.uuid4(), nom="Region X")
    departement = Departement(id=uuid.uuid4(), region_id=region.id, nom="Dep X")
    commune = Commune(id=uuid.uuid4(), departement_id=departement.id, nom="Com X")
    db_session.add_all([region, departement, commune])
    await db_session.commit()

    adhesion = Adhesion(
        id=uuid.uuid4(),
        nom="Doe",
        prenom="John",
        date_naissance=date(1990, 1, 1),
        lieu_naissance="Dakar",
        profession="Dev",
        tel_mobile="770000000",
        tel_fixe=None,
        email=normalize_email("john@example.com"),
        cni="CNI123",
        carte_electeur=None,
        carte_pastef=None,
        region_domicile_id=region.id,
        departement_domicile_id=departement.id,
        commune_domicile_id=commune.id,
        region_militantisme_id=region.id,
        departement_militantisme_id=departement.id,
        commune_militantisme_id=commune.id,
        fonction_professionnelle="Ingénieur",
        engagement=EngagementType.politique,
        commissariat="Com X",
        mode_paiement=PaymentMode.wave,
        montant_adhesion=25000,
        paiement_confirme=False,
        reference_paiement=None,
        certification=True,
        photo_url="/files/photos/x.jpg",
        cv_url="/files/cvs/x.pdf",
        created_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc),
    )
    db_session.add(adhesion)
    await db_session.commit()

    res = await client.get(
        "/api/v1/admin/adhesions/lookup",
        params={"id": str(adhesion.id)},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert res.status_code == 200
    data = res.json()["data"]
    assert data["id"] == str(adhesion.id)
    assert data["email"] == adhesion.email
    assert data["tel_mobile"] == adhesion.tel_mobile
    assert data["region_domicile"]["id"] == str(region.id)


async def test_admin_exports_are_excel_friendly(client, db_session):
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        pytest.skip("openpyxl n'est pas installé")

    repo = UserRepository(db_session)
    user = await repo.create_user(email=normalize_email("admin2@test.com"), password_hash=hash_password("Password123!"))
    await repo.add_role(user_id=user.id, role=AppRole.admin)
    await db_session.commit()

    r = await client.post("/api/v1/auth/login", json={"email": "admin2@test.com", "password": "Password123!"})
    assert r.status_code == 200
    token = r.json()["data"]["accessToken"]

    region = Region(id=uuid.uuid4(), nom="Kédougou")
    departement = Departement(id=uuid.uuid4(), region_id=region.id, nom="Département É")
    commune = Commune(id=uuid.uuid4(), departement_id=departement.id, nom="Commune À")
    db_session.add_all([region, departement, commune])
    await db_session.commit()

    adhesion = Adhesion(
        id=uuid.uuid4(),
        nom="Doe",
        prenom="John",
        date_naissance=date(1990, 1, 1),
        lieu_naissance="Dakar",
        profession="Dev",
        tel_mobile="770000000",
        tel_fixe=None,
        email=normalize_email("john2@example.com"),
        cni="CNI124",
        carte_electeur=None,
        carte_pastef=None,
        region_domicile_id=region.id,
        departement_domicile_id=departement.id,
        commune_domicile_id=commune.id,
        region_militantisme_id=region.id,
        departement_militantisme_id=departement.id,
        commune_militantisme_id=commune.id,
        fonction_professionnelle="Ingénieur",
        engagement=EngagementType.politique,
        commissariat="Com X",
        mode_paiement=PaymentMode.wave,
        montant_adhesion=25000,
        paiement_confirme=False,
        reference_paiement=None,
        certification=True,
        photo_url="/files/photos/x2.jpg",
        cv_url="/files/cvs/x2.pdf",
        created_at=datetime(2026, 1, 2, 3, 4, 5, tzinfo=timezone.utc),
        updated_at=datetime(2026, 1, 2, 3, 4, 5, tzinfo=timezone.utc),
    )
    db_session.add(adhesion)
    await db_session.commit()

    csv_res = await client.get("/api/v1/admin/adhesions/export.csv", headers={"Authorization": f"Bearer {token}"})
    assert csv_res.status_code == 200
    assert csv_res.headers.get("content-disposition") is not None
    content = csv_res.content
    assert content.startswith(b"\xef\xbb\xbf")
    decoded = content.decode("utf-8-sig")
    assert "Kédougou" in decoded

    xlsx_res = await client.get("/api/v1/admin/adhesions/export.xlsx", headers={"Authorization": f"Bearer {token}"})
    assert xlsx_res.status_code == 200
    assert xlsx_res.headers.get("content-type", "").startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert xlsx_res.content[:2] == b"PK"
    wb = load_workbook(BytesIO(xlsx_res.content), read_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "region_domicile_nom" in header
    values = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    assert "Kédougou" in values


async def _login_admin(client, db_session, *, email: str) -> str:
    repo = UserRepository(db_session)
    user = await repo.create_user(email=normalize_email(email), password_hash=hash_password("Password123!"))
    await repo.add_role(user_id=user.id, role=AppRole.admin)
    await db_session.commit()
    r = await client.post("/api/v1/auth/login", json={"email": email, "password": "Password123!"})
    assert r.status_code == 200
    return r.json()["data"]["accessToken"]


async def test_admin_update_status_triggers_email_when_enabled(client, db_session, monkeypatch):
    token = await _login_admin(client, db_session, email="admin3@test.com")

    os.environ["MAIL_ENABLED"] = "true"
    get_settings.cache_clear()

    region = Region(id=uuid.uuid4(), nom="Region X")
    departement = Departement(id=uuid.uuid4(), region_id=region.id, nom="Dep X")
    commune = Commune(id=uuid.uuid4(), departement_id=departement.id, nom="Com X")
    db_session.add_all([region, departement, commune])
    await db_session.commit()

    adhesion = Adhesion(
        id=uuid.uuid4(),
        nom="Doe",
        prenom="John",
        date_naissance=date(1990, 1, 1),
        lieu_naissance="Dakar",
        profession="Dev",
        tel_mobile="770000000",
        tel_fixe=None,
        email=normalize_email("john3@example.com"),
        cni="CNI125",
        carte_electeur=None,
        carte_pastef=None,
        region_domicile_id=region.id,
        departement_domicile_id=departement.id,
        commune_domicile_id=commune.id,
        region_militantisme_id=region.id,
        departement_militantisme_id=departement.id,
        commune_militantisme_id=commune.id,
        fonction_professionnelle="Ingénieur",
        engagement=EngagementType.politique,
        commissariat="Com X",
        mode_paiement=PaymentMode.wave,
        montant_adhesion=25000,
        paiement_confirme=False,
        reference_paiement=None,
        certification=True,
        photo_url="/files/photos/x.jpg",
        cv_url="/files/cvs/x.pdf",
        statut=AdhesionStatus.en_attente,
        created_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc),
    )
    db_session.add(adhesion)
    await db_session.commit()

    calls: list[dict] = []

    def fake_send_email_best_effort(**kwargs):
        calls.append(kwargs)

    monkeypatch.setattr("app.api.v1.routes.admin.send_email_best_effort", fake_send_email_best_effort)

    res = await client.patch(
        f"/api/v1/admin/adhesions/{adhesion.id}",
        json={"statut": "validee"},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert res.status_code == 200
    assert len(calls) == 1
    assert calls[0]["to"] == "john3@example.com"


async def test_admin_confirm_payment_triggers_email_when_enabled(client, db_session, monkeypatch):
    token = await _login_admin(client, db_session, email="admin4@test.com")

    os.environ["MAIL_ENABLED"] = "true"
    get_settings.cache_clear()

    region = Region(id=uuid.uuid4(), nom="Region X")
    departement = Departement(id=uuid.uuid4(), region_id=region.id, nom="Dep X")
    commune = Commune(id=uuid.uuid4(), departement_id=departement.id, nom="Com X")
    db_session.add_all([region, departement, commune])
    await db_session.commit()

    adhesion = Adhesion(
        id=uuid.uuid4(),
        nom="Doe",
        prenom="John",
        date_naissance=date(1990, 1, 1),
        lieu_naissance="Dakar",
        profession="Dev",
        tel_mobile="770000000",
        tel_fixe=None,
        email=normalize_email("john4@example.com"),
        cni="CNI126",
        carte_electeur=None,
        carte_pastef=None,
        region_domicile_id=region.id,
        departement_domicile_id=departement.id,
        commune_domicile_id=commune.id,
        region_militantisme_id=region.id,
        departement_militantisme_id=departement.id,
        commune_militantisme_id=commune.id,
        fonction_professionnelle="Ingénieur",
        engagement=EngagementType.politique,
        commissariat="Com X",
        mode_paiement=PaymentMode.wave,
        montant_adhesion=25000,
        paiement_confirme=False,
        reference_paiement=None,
        certification=True,
        photo_url="/files/photos/x.jpg",
        cv_url="/files/cvs/x.pdf",
        statut=AdhesionStatus.en_attente,
        created_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc),
    )
    db_session.add(adhesion)
    await db_session.commit()

    calls: list[dict] = []

    def fake_send_email_best_effort(**kwargs):
        calls.append(kwargs)

    monkeypatch.setattr("app.api.v1.routes.admin.send_email_best_effort", fake_send_email_best_effort)

    res = await client.patch(
        f"/api/v1/admin/adhesions/{adhesion.id}/payment",
        json={"paiementConfirme": True, "referencePaiement": "REF-1"},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert res.status_code == 200
    assert len(calls) == 1
    assert calls[0]["to"] == "john4@example.com"
